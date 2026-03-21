"""글로벌 업종 매핑표 Excel 생성."""
import json, requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

d = requests.get("http://127.0.0.1:8000/api/market/global-portal", timeout=60).json()
wb = Workbook()

# 스타일
hdr_font = Font(bold=True, color="FFFFFF", size=11)
fills = {
    "us": PatternFill("solid", fgColor="1F4E79"),
    "eu": PatternFill("solid", fgColor="2E75B6"),
    "jp": PatternFill("solid", fgColor="C00000"),
    "kr": PatternFill("solid", fgColor="375623"),
    "L": PatternFill("solid", fgColor="D6E4F0"),
    "M": PatternFill("solid", fgColor="E2EFDA"),
    "S": PatternFill("solid", fgColor="FFF2CC"),
    "hdr": PatternFill("solid", fgColor="404040"),
}
tier_kr = {"L": "대분류", "M": "중분류", "S": "소분류"}
thin = Side(style="thin", color="D9D9D9")
bdr = Border(top=thin, bottom=thin, left=thin, right=thin)
ct = Alignment(horizontal="center", vertical="center")
wr = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_style(cell, fill=None, font=None, align=None):
    cell.border = bdr
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align


# ── Sheet 1: 4개국 비교 ──
ws = wb.active
ws.title = "4개국 비교"
col_defs = [
    ("Q", 4, "hdr"), ("US 업종명", 16, "us"), ("심볼", 8, "us"), ("분류", 6, "us"), ("등락률", 8, "us"),
    ("EU 업종명", 16, "eu"), ("심볼", 10, "eu"), ("분류", 6, "eu"), ("등락률", 8, "eu"),
    ("JP 업종명", 16, "jp"), ("심볼", 8, "jp"), ("분류", 6, "jp"), ("등락률", 8, "jp"),
    ("KR 업종명", 22, "kr"), ("코드", 6, "kr"), ("분류", 6, "kr"), ("등락률", 8, "kr"),
]
for ci, (h, w, fk) in enumerate(col_defs, 1):
    c = ws.cell(row=1, column=ci, value=h)
    set_style(c, fills[fk], hdr_font, ct)
    col_letter = c.column_letter
    ws.column_dimensions[col_letter].width = w

regions_cols = [
    ("us_sectors", [1+0, 2, 3, 4, 5]),
    ("eu_sectors", [0, 6, 7, 8, 9]),
    ("jp_sectors", [0, 10, 11, 12, 13]),
    ("kr_sectors", [0, 14, 15, 16, 17]),
]
max_n = max(len(d.get(rk, [])) for rk, _ in regions_cols)
for ri in range(max_n):
    row = ri + 2
    q_size = (max_n + 3) // 4
    q = ri // q_size + 1
    qc = ws.cell(row=row, column=1, value=f"Q{q}")
    set_style(qc, align=ct)

    for rk, cols in regions_cols:
        items = d.get(rk, [])
        if ri >= len(items):
            for c in cols[1:]:
                set_style(ws.cell(row=row, column=c))
            continue
        it = items[ri]
        tier = it.get("tier", "")
        tf = fills.get(tier)
        c1 = ws.cell(row=row, column=cols[1], value=it.get("name", ""))
        set_style(c1, tf)
        c2 = ws.cell(row=row, column=cols[2], value=it.get("symbol", ""))
        set_style(c2, tf)
        c3 = ws.cell(row=row, column=cols[3], value=tier_kr.get(tier, tier))
        set_style(c3, tf, align=ct)
        c4 = ws.cell(row=row, column=cols[4], value=round(it.get("change_pct", 0) / 100, 4))
        c4.number_format = "+0.00%;-0.00%"
        set_style(c4, tf)


# ── Sheet 2~5: 지역별 상세 ──
region_meta = [
    ("US 미국", "us_sectors", "us"),
    ("EU 유럽", "eu_sectors", "eu"),
    ("JP 일본", "jp_sectors", "jp"),
    ("KR 한국", "kr_sectors", "kr"),
]
for sname, key, fk in region_meta:
    ws2 = wb.create_sheet(sname)
    items = d.get(key, [])
    is_kr = key == "kr_sectors"
    hdrs = ["#", "업종명", "심볼/코드", "분류", "등락률(%)", "상승", "하락"]
    if is_kr:
        hdrs += ["시총(조)", "상위 종목"]
    widths = [4, 22, 12, 8, 10, 6, 6, 10, 45]
    for ci, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        set_style(c, fills[fk], hdr_font, ct)
        ws2.column_dimensions[c.column_letter].width = widths[ci - 1] if ci - 1 < len(widths) else 10

    for ri, it in enumerate(items, 2):
        tier = it.get("tier", "")
        tf = fills.get(tier)
        set_style(ws2.cell(row=ri, column=1, value=ri - 1), tf, align=ct)
        set_style(ws2.cell(row=ri, column=2, value=it.get("name", "")), tf)
        set_style(ws2.cell(row=ri, column=3, value=it.get("symbol", "")), tf)
        set_style(ws2.cell(row=ri, column=4, value=tier_kr.get(tier, tier)), tf, align=ct)
        chg = ws2.cell(row=ri, column=5, value=round(it.get("change_pct", 0) / 100, 4))
        chg.number_format = "+0.00%;-0.00%"
        set_style(chg, tf)
        set_style(ws2.cell(row=ri, column=6, value=it.get("rising", "")), tf, align=ct)
        set_style(ws2.cell(row=ri, column=7, value=it.get("falling", "")), tf, align=ct)
        if is_kr:
            mc = it.get("mktcap", 0)
            set_style(ws2.cell(row=ri, column=8, value=f"{mc / 1e6:.0f}" if mc else ""), tf)
            stocks = it.get("top_stocks", [])
            txt = ", ".join(f'{st["name"]}({st["change_pct"]:+.1f}%)' for st in stocks) if stocks else ""
            set_style(ws2.cell(row=ri, column=9, value=txt), tf, align=wr)


# ── Sheet 6: 크로스 매핑 ──
ws6 = wb.create_sheet("크로스 매핑")
cross_hdrs = ["업종 그룹", "US 미국", "EU 유럽", "JP 일본", "KR 한국"]
for ci, h in enumerate(cross_hdrs, 1):
    c = ws6.cell(row=1, column=ci, value=h)
    set_style(c, fills["hdr"], hdr_font, ct)
ws6.column_dimensions["A"].width = 14
for col in "BCDE":
    ws6.column_dimensions[col].width = 38

cross = [
    ("IT/기술",
     "정보기술, 반도체, 소프트웨어, 사이버보안, 로봇/AI, 클라우드, 반도체장비",
     "기술, 통신",
     "전기기기, 정보통신/서비스, 전기/정밀",
     "핸드셋, 컴퓨터와주변기기, IT서비스, 디스플레이패널, 디스플레이장비및부품, 사무용전자제품, 전자제품"),
    ("금융",
     "금융, 은행, 보험, 증권/브로커, 지방은행",
     "은행, 금융서비스, 보험",
     "은행, 금융(은행제외)",
     "기타금융, 증권, 인터넷과카탈로그소매"),
    ("헬스케어",
     "헬스케어, 바이오테크, 의료기기, 유전체학",
     "헬스케어",
     "의약품",
     "제약, 건강관리장비와용품, 건강관리업체및서비스"),
    ("에너지/화학",
     "에너지, 석유/가스E&P, 우라늄",
     "에너지, 원자재, 화학",
     "에너지자원, 소재/화학, 철강/비철",
     "화학, 석유와가스, 에너지장비및서비스, 철강, 통신장비"),
    ("산업재",
     "산업재, 항공우주/방산, 운송, 항공사, 주택건설",
     "산업재, 자동차, 산업운송, 건설/소재",
     "자동차/운수, 기계, 건설/자재, 운수/물류",
     "건설, 건축자재, 건축제품, 기계, 전기장비, 항공사, 해운사, 도로와철도운송"),
    ("소비재",
     "경기소비재, 소매, 필수소비재",
     "식음료, 소매, 개인/가정용품, 여행/레저",
     "식품, 상사/도소매",
     "식품, 음료, 식품과기본식료품소매, 백화점과일반상점, 전문소매, 섬유의류, 화장품"),
    ("소재/광업",
     "소재, 금속/광업, 리튬/배터리, 구리광업, 희토류, 임업/목재",
     "미디어",
     "-",
     "포장재, 가정용기기와용품, 가구, 종이와목재"),
    ("유틸/통신",
     "유틸리티, 태양광, 클린에너지, 커뮤니케이션",
     "유틸리티",
     "전력/가스",
     "전기유틸리티, 가스유틸리티, 복합유틸리티, 다각화된통신서비스, 무선통신서비스"),
    ("기타",
     "부동산",
     "부동산",
     "부동산, 수산/농림",
     "부동산, 레저용장비와제품, 광고, 출판, 상업서비스, 다각화된소비자서비스, 복합기업, 무역회사와판매업체"),
]
for ri, (grp, us, eu, jp, kr) in enumerate(cross, 2):
    set_style(ws6.cell(row=ri, column=1, value=grp), font=Font(bold=True))
    set_style(ws6.cell(row=ri, column=2, value=us), align=wr)
    set_style(ws6.cell(row=ri, column=3, value=eu), align=wr)
    set_style(ws6.cell(row=ri, column=4, value=jp), align=wr)
    set_style(ws6.cell(row=ri, column=5, value=kr), align=wr)

path = r"C:\trade-system\output\market_mapping.xlsx"
wb.save(path)
print(f"Saved: {path}")
print(f"Sheets: {wb.sheetnames}")
us = len(d.get("us_sectors", []))
eu = len(d.get("eu_sectors", []))
jp = len(d.get("jp_sectors", []))
kr = len(d.get("kr_sectors", []))
print(f"US:{us} EU:{eu} JP:{jp} KR:{kr} Total:{us+eu+jp+kr}")
